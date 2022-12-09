import pandas as pd
import warnings
from datetime import datetime
from dateutil.relativedelta import relativedelta

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, colors
from openpyxl.styles.colors import Color, ColorDescriptor
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


warnings.filterwarnings(action="ignore",)
df = pd.read_excel(r'F:\PROJECTS\JK_SOFT\GIT IMMILYTICS\morning_star\Source Data\Open Process Data - Morningstar.xlsx')



def add_designs(no_shts,file_path,doc_type='',date_months=''):
    
    exp_month = ((datetime.today())) + relativedelta(days=+240) 
    this_month = ((datetime.today()))
    next_month = ((datetime.today())) + relativedelta(months=+1) 

    end_month_str = pd.to_datetime(exp_month)
    this_month_str = pd.to_datetime(this_month)
    next_month_str = pd.to_datetime(next_month).date()



    # print ('\nAdding designs to the processed file..')
    book = load_workbook(file_path)

    for sheet in range(no_shts):
        ws = book[book.sheetnames[sheet]]
        if ws:
            rows = ws.max_row 
            cols= ws.max_column 

            if rows<2:
                rows = 2
                ws.cell(row=2, column = 1).value = "No Records Found"

            if sheet == 0:   
                ws.freeze_panes = ws['D2']
            else:
                ws.freeze_panes = ws['F2']

            for y in range(rows):
                for z in range(cols):

                    ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri (Body)', size = 11)

                    ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal="left", vertical="bottom")

                    ws.cell(row=y+1, column=z+1).font= Font(name = 'Calibri (Body)', size= 11)

                    if (sheet== 4 or sheet== 5) and  (doc_type == 'comcast'):  #
                        if str(ws.cell(row=1, column=z+1).value) in date_months:
                            doc_date = ws.cell(row=y+1, column=z+1).value

                        #old working conditions might be used later:
                            # try:
                            #     if (doc_date>=this_month) and (doc_date<=exp_month):
                                    
                            #         if(doc_date.month == this_month.month) and (doc_date.year == this_month.year):
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid") #red    
                            #         elif(doc_date.month == next_month.month) and (doc_date.year == next_month.year):
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type = "solid") #orange
                            #         else:
                            #             ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type = "solid") #yellow
                            # except:
                            #     pass

                        # coloring blank cell condition:

                        #     if str(doc_date) in ['','nan','NaT','NaN','Nan','None']:
                        #         ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid") #red
                            
                            # #new conditions:
                            try:
                                if (doc_date<=exp_month):
                                    
                                    if(doc_date <= this_month):
                                        ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid") #red    
                                    else:
                                        ws.cell(row=y+1, column=z+1).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type = "solid") #yellow
                            except:
                                pass

                    ws.cell(row=y+1, column=z+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    if y == 0:
                        ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri',size = 12, color = 'FFFFFF')

                        ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal="center", vertical="center")

        for cl in range(cols):
            if cl <= cols:
                ws.column_dimensions[get_column_letter(cl+1)].width = 15

        for rw in range(rows+1):
            if rw <= rows:
                ws.row_dimensions[rw].height = 30

        table = Table(displayName="Table{}".format(sheet+1), ref="A1:" + get_column_letter(cols) + str(rows))

        #     ws.cell(row=2, column = 1).value = "No Records Found"

        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        table.tableStyleInfo = style
        ws.add_table(table)

        book.save(file_path)
    book.close()
        

df = df.astype(str)

for column in df.columns:
    for index,row in df.iterrows():
        try:
            df[column][index] = pd.to_datetime(str(df[column][index]),dayfirst=True)
            df[column][index] = datetime.strftime(df[column][index],format='%m/%d/%Y')
        except:
            pass

writer = pd.ExcelWriter('qqq.xlsx', engine='xlsxwriter', date_format='m/d/yyyy')
df.to_excel(writer, "Internal-PERM Report", startrow=0, index=False)
writer.save()
add_designs(no_shts=1,file_path='qqq.xlsx',doc_type='',date_months='')

# print(df)


# for index,row in df.iterrows():
#     print(row[0]) 


#     quit()
# quit()
# df['comment'] = ''
# # print(df)
# # df['comment'][index] = "he's female"

# for index,column in enumerate(df['GENDER']):
#     print(index,column)
#     # if index!=5:
#         # if df['GENDER'][index] == 'FEMALE':   
#         #     df['comment'][index] = "he's female"
            

#         # else:
#         #     df['comment'][index] = "he's male"
        
#         # try:
#         #     # print(type(df['DATE'][index]))
#         #     df['DATE'][index] = pd.to_datetime(df['DATE'][index],format="%m/%d/%Y")
#         #     df['DATE'][index] =  datetime.strftime(df['DATE'][index],"%Y-%m-%d")
            
#         #     # print('hi')
#         #     # print(type(df['DATE'][index]))
#         # except:
#         #     pass


# df.to_excel('res.xlsx')
# # df.loc[df["GENDER"] == 'FEMALE', "comment"] = 'whamen'
# print(df)

# # df['review'] = ''

# # for index,row in df.iterrows():
# #     if row['MARK']<20:
# #         df['review'][index] = 'poda punda'
# #     elif row['MARK']<40:
# #         df['review'][index] = 'nottu'
# # # print(df[df['GENDER'] == 'FEMALE']) 
# # df.to_csv('Book1.csv')

# # df.loc[df['comment']=="whamen","review"] = 'sappu'