import numpy as np
import pandas as pd
import os
import glob
from datetime import *
from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

cwd = os.path.dirname(os.path.realpath(__file__))
print(cwd)
os.chdir(cwd)

def process_beneficiary(df_ben):

    source_excel_headers_actemp = ['Case No',"Beneficiary Name","Birth Country","Citizenship","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","I-129S Expires","Petition Expiration Date PED","EAD Expiration","Visa Priority Date","Visa Preference","Visa Country of Chargeability","Visa Priority Note","Management Info Employee ID","Management Info Job Title","Management Info Job Location City","Management Info Job Location State","Current Process Type","Management Info Business Partner Name"] 

    result_excel_headers_actemp = ['Unique Record Id (BT)',"Employee Name","Country of Birth","Country of Citizenship","Current Status","Current Status Expiration Date","I-797 Expiration Date","I-94 Status","I-94 Expiration Date","NIV Max Out Date","I-129S Expiration Date","PED","EAD Expiration Date","Visa Priority Date","Visa Preference","Visa Country of Chargeability","Visa Priority Note","Employee Id","Job Title","Work Location City","Work Location State","Current Case Type","HRBP"]


    df_tab1 = pd.DataFrame()

    for x,y in zip(source_excel_headers_actemp,result_excel_headers_actemp):     
        df_tab1[y] = df_ben[x]

    for index,row in df_tab1.iterrows():
        if type(row['Visa Priority Date']) == str:
            df_tab1['Visa Priority Date'][index] = (row['Visa Priority Date'].split(' ')[1].strip())
            # df_tab1['Visa Priority Date'][index] = datetime.strptime(df_tab1['Visa Priority Date'][index],'%m/%d/%Y')
        else:
            df_tab1['Visa Priority Date'][index] = None

    for index,row in df_tab1.iterrows() :
        for column in result_excel_headers_actemp:
            if 'D/S' in str(row[column]):
                (df_tab1[column][index]) = None

    for y in (result_excel_headers_actemp):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
            if y != 'Visa Priority Date': 
                #(this above col has string data already so no conv needed)

                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab1[y] = pd.to_datetime(df_tab1[y], format='%Y-%m-%d',errors='coerce').dt.date
  
    return df_tab1
   
def process_casefile(df_case,src_name,df_tab1):

    file_gen_date =  date.today().strftime("%m%d%y")
    file_path = "Processed Reports Folder/{}_StatusReport_{}.xlsx".format(src_name, file_gen_date)
    

    source_excel_headers_niv = ['Case No',"Petitioner","Beneficiary Name","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","Petition Expiration Date PED","EAD Expiration","HR Info - Department","Management Info Employee ID","Management Info Job Title","Process Case No","Case Opened","Process Type","Process Reference","Application Filed","Final Action Status","Final Action Date","Summary Case Disposition"]

    result_excel_headers_niv = ['Unique Record Id (BT)',"Petitioner","Beneficiary Name","Current Status","Current Status Expiration Date","I-797 Expiration Date","Entry Status","I-94 Expiration Date","NIV Max Out Date","PED ","EAD Expiration Date","HRBP","Employee Id","Job Title","Case Id","Case Opened Date","Case Type","Case Reference","Case Filed Date ","Final Action Status","Final Action Date","Summary Case Disposition"]

    date_columns = []

    df_tab2 = pd.DataFrame()

    for x,y in zip(source_excel_headers_niv,result_excel_headers_niv):     
        df_tab2[y] = df_case[x]

    for index,row in df_tab2.iterrows() :
        for column in result_excel_headers_niv:
            if 'D/S' in str(row[column]):
                (df_tab2[column][index]) = None

    for y in (result_excel_headers_niv):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab2[y] = pd.to_datetime(df_tab2[y], format='%Y-%m-%d',errors='coerce').dt.date
   
    df_tab2 = df_tab2[(df_tab2['Case Type'] == 'H-1B Professional') |  
                      (df_tab2['Case Type'] == 'L-1A Intracompany Transfer') | 
                      (df_tab2['Case Type'] == 'L-1B Intracompany Transfer') |
                      (df_tab2['Case Type'] == 'E-3 Treaty Professional')    |
                      (df_tab2['Case Type'] == 'L-1A/B Intracompany Transfer') | 
                      (df_tab2['Case Type'] == 'TN Extension') |
                      (df_tab2['Case Type'] == 'L Blanket') |
                      (df_tab2['Case Type'] == 'H-4 Derivative') ]

    source_excel_headers_perm = ['Case No',"Petitioner","Beneficiary Name","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","Petition Expiration Date PED","EAD Expiration","HR Info - Department","HR Info - Department Number","HR Info - Job Code","Management Info Employee ID","Management Info Job Title","Process Case No","Case Opened","Process Type","Process Reference","LC First Filing Date","LC Last Filing Date","Application Filed","Final Action Status","Final Action Date","Summary Case Disposition"] 

    result_excel_headers_perm = ['Unique Record Id (BT)',"Petitioner","Beneficiary Name","Current Status","Current Status Expiration Date","I-797 Expiration Date","Entry Status","I-94 Expiration Date","NIV Max Out Date","PED ","EAD Expiration Date","HRBP","REQ #","PERM Job Title","Employee Id","Job Title","Case Id","Case Opened Date","Case Type","Case Reference","LC First Filing Date","LC Last Filing Date","Case Filed Date ","Final Action Status","Final Action Date","Summary Case Disposition"]


    df_tab3 = pd.DataFrame()

    for x,y in zip(source_excel_headers_perm,result_excel_headers_perm):     
        df_tab3[y] = df_case[x]

    for index,row in df_tab3.iterrows() :
        for column in result_excel_headers_perm:
            if 'D/S' in str(row[column]):
                (df_tab3[column][index]) = None

    for y in (result_excel_headers_perm):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab3[y] = pd.to_datetime(df_tab3[y], format='%Y-%m-%d',errors='coerce').dt.date

        

    df_tab3 = df_tab3[(df_tab3['Case Type'] == 'Labor Cert PERM')]


    source_excel_headers_pr = ['Case No',"Petitioner","Beneficiary Name","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","Petition Expiration Date PED","EAD Expiration","HR Info - Department","Management Info Employee ID","Management Info Job Title","Process Case No","Case Opened","Process Type","Process Reference","I-140 filing deadline","Application Filed","Final Action Status","Final Action Date","Summary Case Disposition"] 

    result_excel_headers_pr = ['Unique Record Id (BT)',"Petitioner","Beneficiary Name","Current Status","Current Status Expiration Date","I-797 Expiration Date","Entry Status","I-94 Expiration Date","NIV Max Out Date","PED ","EAD Expiration Date","HRBP","Employee Id","Job Title","Case Id","Case Opened Date","Case Type","Case Reference","I-140 filing deadline","Case Filed Date ","Final Action Status","Final Action Date","Summary Case Disposition"]


    df_tab4 = pd.DataFrame()

    for x,y in zip(source_excel_headers_pr,result_excel_headers_pr):     
        df_tab4[y] = df_case[x]

    for index,row in df_tab4.iterrows() :
        for column in result_excel_headers_pr:
            if 'D/S' in str(row[column]):
                (df_tab4[column][index]) = None

    for y in (result_excel_headers_pr):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab4[y] = pd.to_datetime(df_tab4[y], format='%Y-%m-%d',errors='coerce').dt.date


    df_tab4 = df_tab4[(df_tab4['Case Type'] == 'I-140 LC Required') |
                      (df_tab4['Case Type'] == 'I-140 LC Exempt') |
                      (df_tab4['Case Type'] == 'AOS Employment')]


    source_excel_headers_h1b = ['Case No',"Petitioner","Beneficiary Name","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","Petition Expiration Date PED","EAD Expiration","HR Info - Department","Management Info Employee ID","Management Info Job Title","Process Case No","Case Opened","Process Type","Process Reference","Application Filed","Final Action Status","Final Action Date","Summary Case Disposition"] 

    result_excel_headers_h1b = ['Unique Record Id (BT)',"Petitioner","Beneficiary Name","Current Status","Current Status Expiration Date","I-797 Expiration Date","Entry Status","I-94 Expiration Date","NIV Max Out Date","PED ","EAD Expiration Date","HRBP","Employee Id","Job Title","Case Id","Case Opened Date","Case Type","Case Reference","Case Filed Date ","Final Action Status","Final Action Date","Summary Case Disposition"]


    df_tab5 = pd.DataFrame()

    for x,y in zip(source_excel_headers_h1b,result_excel_headers_h1b):     
        df_tab5[y] = df_case[x]

    for index,row in df_tab5.iterrows() :
        for column in result_excel_headers_h1b:
            if 'D/S' in str(row[column]):
                (df_tab5[column][index]) = None

    for y in (result_excel_headers_h1b):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab5[y] = pd.to_datetime(df_tab5[y], format='%Y-%m-%d',errors='coerce').dt.date

    df_tab5 = df_tab5[(df_tab5['Case Type'] == 'H-1B CAP')]

# tab6 reading from the new excel file
    Approved_cases_file = glob.glob("Source Data/*Approved*")

    df_approved = pd.read_excel(Approved_cases_file[0]) 
    df_tab6 = pd.DataFrame()


    source_excel_headers_aprov = ['Case No',"Petitioner","Beneficiary Name","Current Status","Current Status Expires","I-797 Expires","Entry Status","I94 Expires","NIV Max Out Date","Petition Expiration Date PED","EAD Expiration","HR Info - Department","Management Info Employee ID","Management Info Job Title","Process Case No","Case Opened","Process Type","Process Reference","Application Filed","Final Action Status","Final Action Date","Summary Case Disposition"] 

    result_excel_headers_aprov = ['Unique Record Id (BT)',"Petitioner","Beneficiary Name","Current Status","Current Status Expiration Date","I-797 Expiration Date","Entry Status","I-94 Expiration Date","NIV Max Out Date","PED ","EAD Expiration Date","HRBP","Employee Id","Job Title","Case Id","Case Opened Date","Case Type","Case Reference","Case Filed Date ","Final Action Status","Final Action Date","Summary Case Disposition"]

    for x,y in zip(source_excel_headers_aprov,result_excel_headers_aprov):     
        df_tab6[y] = df_approved[x]

    for index,row in df_tab6.iterrows() :
        for column in result_excel_headers_aprov:
            if 'D/S' in str(row[column]):
                (df_tab6[column][index]) = None

    for y in (result_excel_headers_aprov):     
        if 'Date' in y or 'PED' in y or 'dead' in y:
                # if "1900-01-01" in df_tab1[y]:
                #     df_tab1[y] = ""
                # else:
                df_tab6[y] = pd.to_datetime(df_tab6[y], format='%Y-%m-%d',errors='coerce').dt.date

    df_tab6 = df_tab6[(df_tab6['Final Action Status'] == 'Approved') |
                      (df_tab6['Final Action Status'] == 'PERM Certified') |
                      (df_tab6['Final Action Status'] == 'Granted')]
    
    df_tab1 = df_tab1.sort_values(by='Employee Name',ascending=True)
    df_tab2 = df_tab2.sort_values(by='Beneficiary Name',ascending=True)
    df_tab3 = df_tab3.sort_values(by='Beneficiary Name',ascending=True)
    df_tab4 = df_tab4.sort_values(by='Beneficiary Name',ascending=True)
    df_tab5 = df_tab5.sort_values(by='Beneficiary Name',ascending=True)
    df_tab6 = df_tab6.sort_values(by='Beneficiary Name',ascending=True)


    writer = pd.ExcelWriter(file_path, engine = 'xlsxwriter',date_format='m/d/yyyy')
    df_tab1.to_excel(writer,'Active Employees List',index=False)
    df_tab2.to_excel(writer,'NIV Cases', index=False)
    df_tab3.to_excel(writer,'PERM Cases', index=False)
    df_tab4.to_excel(writer,'PR Cases', index=False)
    df_tab5.to_excel(writer,'H-1B Cap Cases', index=False)
    df_tab6.to_excel(writer,'Recently Approved Cases', index=False)
    
    writer.save()
    # writer.close()

    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine = 'openpyxl')
    writer.book = book

    for x in range(6):
        ws = book[book.sheetnames[x]]
        if ws:
            rows = ws.max_row 
            cols= ws.max_column 

            if x == 0:   
                ws.freeze_panes = ws['D2']
            else:
                ws.freeze_panes = ws['F2']

            for y in range(rows):
                for z in range(cols):

                    ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri (Body)', size = 11)

                    ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal="center", vertical="justify")

                    ws.cell(row=y+1, column=z+1).font= Font(name = 'Calibri (Body)', size= 11)

                    ws.cell(row=y+1, column=z+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    if y == 0:

                        ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri',size = 12, color = 'FFFFFF')

        for cl in range(cols):
            if cl <= cols:
                ws.column_dimensions[get_column_letter(cl+1)].width = 15

        for rw in range(rows+1):
            if rw <= rows:
                ws.row_dimensions[rw].height = 30

        table = Table(displayName="Table{}".format(x+1), ref="A1:" + get_column_letter(cols) + str(rows))

        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # for z in range(cols):
        #    ws.cell(row=1, column=z+1).font = Font(size = 12,color = 'ffffff')

    writer.save()
    writer.close()
    
    
def start():
    for name in glob.glob('Source Data/Active Beneficiary*'):
        beneficiary_file = os.path.abspath(name)
        src_name = os.path.basename(name)
        print ('Processing..  ' + str(beneficiary_file))
        df_ben = pd.read_excel(beneficiary_file)
        src_name = src_name.split('-')[1].strip()
        src_name= os.path.splitext(src_name)[0]
        print(src_name)

        # quit()
        df_tab1 = process_beneficiary(df_ben)
        print('Processed Successfully beneficiary file\n\n')


    for name in glob.glob('Source Data/Open process Data*'):
            case_file = os.path.abspath(name)
            src_name = os.path.basename(name)
            src_name = src_name.split('-')[1].strip()
            src_name= os.path.splitext(src_name)[0]
            print(src_name)
            # quit()
            print ('processing..  ' + str(case_file))
            df_case = pd.read_excel(case_file)
            process_casefile(df_case,src_name,df_tab1)
            print('Processed Successfully case file')


print('\n\nProgram Execution Started..\n')
start()
print('\n\nFinished..')


